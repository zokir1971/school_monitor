from __future__ import annotations

from openpyxl import load_workbook


def normalize_text(value: object | None) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def read_schools_from_xlsx(file_path: str) -> list[dict]:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            return []

        headers = [normalize_text(col) for col in rows[0]]

        def get_col_index(name: str) -> int:
            if name not in headers:
                raise ValueError(f"В Excel-файле нет колонки: {name}")
            return headers.index(name)

        idx_region = get_col_index("Облыс")
        idx_district = get_col_index("Аудан")
        idx_settlement = get_col_index("Елді мекен")
        idx_school = get_col_index("Ұйым атауы")

        idx_address = None
        if "Қазақ тіліндегі заңды мекен-жайы" in headers:
            idx_address = headers.index("Қазақ тіліндегі заңды мекен-жайы")

        items: list[dict] = []

        for row in rows[1:]:
            region_name = normalize_text(row[idx_region] if idx_region < len(row) else None)
            district_name = normalize_text(row[idx_district] if idx_district < len(row) else None)
            settlement = normalize_text(row[idx_settlement] if idx_settlement < len(row) else None)
            school_name = normalize_text(row[idx_school] if idx_school < len(row) else None)

            address = None
            if idx_address is not None and idx_address < len(row):
                address = normalize_text(row[idx_address]) or None

            if not region_name or not district_name or not school_name:
                continue

            items.append(
                {
                    "region_name": region_name,
                    "district_name": district_name,
                    "settlement": settlement or None,
                    "school_name": school_name,
                    "address": address,
                }
            )

        return items
    finally:
        wb.close()
